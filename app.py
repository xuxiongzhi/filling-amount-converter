import streamlit as st
import pandas as pd
import re
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import traceback
import shutil
import tempfile
from PIL import Image as PILImage

# 定义 log_message 函数
def log_message(message):
    """记录日志到 session_state"""
    if 'logs' not in st.session_state:
        st.session_state.logs = []
    st.session_state.logs.append(message)

# 初始化 session_state
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'output_path' not in st.session_state:
    st.session_state.output_path = None
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'initialized' not in st.session_state:
    st.session_state.initialized = False

# 调试环境信息（仅初始化时记录）
if not st.session_state.initialized:
    log_message(f"Python 版本: {sys.version}")
    log_message(f"Streamlit 版本: {st.__version__}")
    log_message(f"Pandas 版本: {pd.__version__}")
    try:
        import numpy
        log_message(f"Numpy 版本: {numpy.__version__}")
    except ImportError:
        log_message("Numpy 未安装")
    try:
        import pyarrow
        log_message(f"Pyarrow 版本: {pyarrow.__version__}")
    except ImportError:
        log_message("Pyarrow 未安装")
    st.session_state.initialized = True

# 预编译正则表达式
SIZE_REGEX = re.compile(r'^(XXS|XS|S|M|L|XL|2XL|3XL|4XL|5XL)$', re.IGNORECASE)
PIECE_NAME_REGEX = re.compile(r'裁片名\s*[:：]\s*(\S+)')
NUMERIC_INDEX_REGEX = re.compile(r'^\d+$')
FILLING_AMOUNT_REGEX = re.compile(r'^\d*\.?\d+$')

def make_unique_sheet_title(workbook, desired_title_base):
    """确保工作表名称唯一，符合 Excel 规范"""
    sanitized_title_base = re.sub(r'[\\/*?:\[\]]', '_', desired_title_base)[:31]
    if not sanitized_title_base:
        sanitized_title_base = "Sheet"
    if sanitized_title_base not in workbook.sheetnames:
        return sanitized_title_base
    base_for_suffix = sanitized_title_base[:28]
    count = 1
    while True:
        new_title = f"{base_for_suffix}_{count}"
        if new_title not in workbook.sheetnames:
            return new_title
        count += 1
        if count > 999:
            return f"{base_for_suffix}_fallback_{os.urandom(4).hex()}"

def extract_images_from_sheet_object(openpyxl_sheet_obj):
    """从工作表提取图片数据"""
    images_data = []
    if hasattr(openpyxl_sheet_obj, '_images') and openpyxl_sheet_obj._images:
        for img in openpyxl_sheet_obj._images:
            try:
                # openpyxl 的 Image 对象没有 data 属性，使用 PIL 读取图片
                if hasattr(img, 'ref') and img.ref:
                    with PILImage.open(img.ref) as pil_img:
                        # 转换为 PNG 格式保存到 BytesIO
                        img_bytes = BytesIO()
                        pil_img.save(img_bytes, format='PNG')
                        images_data.append({
                            'data': img_bytes,
                            'width': img.width if hasattr(img, 'width') else pil_img.width,
                            'height': img.height if hasattr(img, 'height') else pil_img.height
                        })
                else:
                    log_message(f"图片缺少 ref 属性，无法提取")
            except Exception as img_err:
                log_message(f"提取图片时出错: {img_err}")
    return images_data

def extract_data_from_dataframe(df):
    """从 DataFrame 提取充绒量数据"""
    piece_name = "未命名裁片"
    data = {}
    sizes = set()
    max_index = 0
    current_size = None
    header_found = False
    filling_col = None

    for _, row_series in df.head(10).iterrows():
        row_values = [str(x) if pd.notna(x) else "" for x in row_series.values]
        row_str = ' '.join(row_values)
        match = PIECE_NAME_REGEX.search(row_str)
        if match:
            piece_name = match.group(1)
            break

    for _, row_series in df.iterrows():
        row_values = [str(x).strip() if pd.notna(x) else "" for x in row_series.values]
        if not any(row_values):
            continue
        if not header_found and '规格' in row_values and '单片充绒量' in row_values:
            header_found = True
            try:
                filling_col = row_values.index('单片充绒量')
            except ValueError:
                log_message(f"表头中未找到'单片充绒量'，裁片 '{piece_name}' 无法处理")
                return {}, piece_name, [], 0
            continue
        if header_found and row_values and row_values[0] and SIZE_REGEX.match(row_values[0]):
            current_size = row_values[0].upper()
            sizes.add(current_size)
        if current_size:
            try:
                index = None
                for cell_val in row_values:
                    if NUMERIC_INDEX_REGEX.match(cell_val):
                        index = int(cell_val)
                        max_index = max(max_index, index)
                        break
                if index is not None and filling_col is not None and filling_col < len(row_values):
                    potential_filling = row_values[filling_col]
                    filling_amount = ""
                    if potential_filling and potential_filling.lower() != "nan" and FILLING_AMOUNT_REGEX.match(potential_filling):
                        try:
                            filling_amount = float(potential_filling)
                        except ValueError:
                            pass
                    if index is not None:
                        if current_size not in data:
                            data[current_size] = {}
                        data[current_size][index] = filling_amount
            except Exception as e:
                log_message(f"处理数据行出错 (裁片: {piece_name}): {e}")

    sorted_sizes = sorted(list(sizes), key=lambda x: (
        -100 if x == 'XXS' else -50 if x == 'XS' else 0 if x == 'S' else
        10 if x == 'M' else 20 if x == 'L' else 30 if x == 'XL' else
        (int(x[0]) * 10 + 40) if x[0].isdigit() and x.upper().endswith('XL') else 100
    ))
    return data, piece_name, sorted_sizes, max_index

def populate_output_sheet(worksheet, data, piece_name, sizes, max_index, images_to_add=None):
    """填充工作表数据和图片"""
    for size in sizes:
        data.setdefault(size, {})
    headers = [''] + sizes
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    current_row = 2
    for idx in range(1, max_index + 1):
        cell = worksheet.cell(row=current_row, column=1, value=f"{piece_name}{idx}充绒")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_offset, size in enumerate(sizes):
            value = data.get(size, {}).get(idx, '')
            cell = worksheet.cell(row=current_row, column=col_offset + 2, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        worksheet.column_dimensions[column_letter].width = max((max_length + 2) * 1.2, 8)

    if images_to_add:
        image_start_row = worksheet.max_row + 3
        current_row = image_start_row
        for img_info in images_to_add:
            try:
                img = OpenpyxlImage(img_info['data'])
                if img_info['width'] and img_info['height']:
                    img.width = img_info['width']
                    img.height = img_info['height']
                worksheet.add_image(img, f'A{current_row}')
                rows_for_image = (img_info['height'] // 20 if img_info['height'] and img_info['height'] > 0 else 10) + 3
                current_row += max(5, rows_for_image)
            except Exception as e:
                log_message(f"添加图片到工作表 '{worksheet.title}' 出错: {e}")

def process_file(input_path, output_path):
    """处理 Excel 文件"""
    if not os.path.exists(input_path):
        log_message(f"输入文件 '{input_path}' 不存在")
        return False

    if not input_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        log_message(f"输入文件 '{input_path}' 不是 Excel 文件，可能无法正确处理")
        return False

    try:
        shutil.copy2(input_path, output_path)
        log_message(f"文件已复制到: {output_path}")
    except Exception as e:
        log_message(f"复制文件出错: {e}")
        return False

    try:
        modified_workbook = load_workbook(output_path)
        original_sheet_names = modified_workbook.sheetnames
        log_message(f"已打开文件，包含工作表: {original_sheet_names}")
    except Exception as e:
        log_message(f"打开文件失败: {e}")
        return False

    any_sheet_transformed = False
    for sheet_name in original_sheet_names:
        log_message(f"\n处理工作表: {sheet_name}")
        ws = modified_workbook[sheet_name]
        try:
            df = pd.read_excel(input_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        except Exception as e:
            log_message(f"无法读取工作表 '{sheet_name}' 数据: {e}")
            continue

        data, piece_name, sorted_sizes, max_index = extract_data_from_dataframe(df)
        if data and sorted_sizes and max_index > 0:
            any_sheet_transformed = True
            images = extract_images_from_sheet_object(ws)
            log_message(f"工作表 '{sheet_name}' 提取到 {len(images)} 张图片")
            if ws.merged_cells.ranges:
                for merged_range in list(ws.merged_cells.ranges):
                    try:
                        ws.unmerge_cells(str(merged_range))
                    except Exception as e:
                        log_message(f"解除合并单元格 {merged_range} 失败: {e}")
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            new_title = make_unique_sheet_title(modified_workbook, piece_name if piece_name != "未命名裁片" else sheet_name)
            if ws.title != new_title:
                ws.title = new_title
            populate_output_sheet(ws, data, piece_name, sorted_sizes, max_index, images_to_add=images)
            log_message(f"工作表 '{new_title}' 已更新")
        else:
            log_message(f"工作表 '{sheet_name}' 数据格式不符合要求，保持原样")

    if any_sheet_transformed:
        log_message("至少有一个工作表被转换")
    else:
        log_message("没有工作表被转换，输出文件为原始副本")

    try:
        modified_workbook.save(output_path)
        log_message(f"处理完成，输出文件: {output_path}")
        return True
    except Exception as e:
        log_message(f"保存文件失败: {e}")
        traceback.print_exc()
        return False

# Streamlit 界面
st.title("充绒量数据格式转换工具")
st.markdown("上传 Excel 文件，转换充绒量数据格式后下载结果。支持多工作表和图片保留。")
st.info("请上传 .xlsx, .xls 或 .xlsm 文件，最大 50MB。处理完成后点击下载按钮获取结果。")

# 日志显示
log_container = st.container()
with log_container:
    st.subheader("处理日志")
    st.text_area("日志", value="\n".join(st.session_state.logs), height=200, key="log_area", disabled=True)

uploaded_file = st.file_uploader("选择 Excel 文件", type=['xlsx', 'xls', 'xlsm'], key="file_uploader")

# 处理上传的文件
if uploaded_file and not st.session_state.processing:
    st.session_state.logs = []  # 清空日志
    st.session_state.processing = True
    st.session_state.output_path = None  # 重置输出路径
    with st.spinner("正在处理文件..."):
        try:
            # 创建持久临时文件
            temp_dir = tempfile.mkdtemp()
            input_path = os.path.join(temp_dir, uploaded_file.name)
            output_path = os.path.join(temp_dir, f"{os.path.splitext(uploaded_file.name)[0]}_转换后.xlsx")
            
            # 保存上传文件
            with open(input_path, 'wb') as f:
                f.write(uploaded_file.read())
            
            # 检查文件大小（限制 50MB）
            max_size_mb = 50
            if os.path.getsize(input_path) > max_size_mb * 1024 * 1024:
                log_message(f"文件大小超过 {max_size_mb}MB 限制")
                st.error(f"文件大小超过 {max_size_mb}MB 限制")
            else:
                # 处理文件
                if process_file(input_path, output_path):
                    st.session_state.output_path = output_path
                    log_message("文件处理成功，请点击下方按钮下载结果")
                    st.success("文件处理完成！")
                else:
                    st.error("文件处理失败，请检查日志")
        except Exception as e:
            log_message(f"处理过程中发生错误: {e}")
            st.error(f"处理失败: {e}")
            traceback.print_exc()
        finally:
            st.session_state.processing = False

# 下载按钮
if st.session_state.output_path:
    log_message(f"检查下载按钮: output_path={st.session_state.output_path}, exists={os.path.exists(st.session_state.output_path)}")
    if os.path.exists(st.session_state.output_path):
        with open(st.session_state.output_path, 'rb') as f:
            st.download_button(
                label="下载处理结果",
                data=f,
                file_name=os.path.basename(st.session_state.output_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=st.session_state.processing
            )
    else:
        log_message("输出文件不存在，无法显示下载按钮")
        st.error("输出文件已失效，请重新上传文件")
